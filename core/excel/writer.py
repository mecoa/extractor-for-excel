from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from models.extract_result import ExtractResult
from models.field import Confidence


CONFIDENCE_FILLS = {
    Confidence.HIGH: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    Confidence.MEDIUM: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    Confidence.LOW: PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    Confidence.MISSING: PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


class ExcelWriter:
    def __init__(self, template_path: str, output_path: str):
        self.template_path = template_path
        self.output_path = output_path

    def write_results(
        self,
        results: list[ExtractResult],
        field_names: list[str],
    ):
        wb = load_workbook(self.template_path)
        ws = wb.active
        headers = [str(cell.value or "") for cell in ws[1]]
        col_indices = []
        for fname in field_names:
            try:
                col_indices.append(headers.index(fname) + 1)
            except ValueError:
                col_indices.append(-1)

        for result in results:
            row_num = result.row_index + 2
            for fname, col in zip(field_names, col_indices):
                if col < 1:
                    continue
                fr = result.fields.get(fname)
                if fr is None:
                    continue
                cell = ws.cell(row=row_num, column=col, value=fr.display_value)
                fill = CONFIDENCE_FILLS.get(fr.confidence)
                if fill:
                    cell.fill = fill

        wb.save(self.output_path)
